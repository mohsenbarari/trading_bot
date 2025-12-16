# bot/handlers/trade_history.py
"""هندلرهای تاریخچه معاملات"""

from aiogram import Router, types, F, Bot
from aiogram.fsm.context import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.exceptions import TelegramBadRequest
from sqlalchemy import select, and_, or_
from sqlalchemy.orm import joinedload
from typing import Optional
from datetime import datetime, timedelta
import os
import tempfile

from models.user import User
from models.trade import Trade, TradeType, TradeStatus
from models.offer import Offer, OfferType
from models.commodity import Commodity
from core.db import AsyncSessionLocal
import jdatetime
from datetime import timezone, timedelta

# تایم‌زون ایران (UTC+3:30)
IRAN_TZ = timezone(timedelta(hours=3, minutes=30))

router = Router()


def get_trade_history_keyboard(target_user_id: int) -> InlineKeyboardMarkup:
    """کیبورد تاریخچه معاملات"""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📥 دانلود Excel", callback_data=f"export_excel_{target_user_id}"),
            InlineKeyboardButton(text="📄 دانلود PDF", callback_data=f"export_pdf_{target_user_id}")
        ],
        [
            InlineKeyboardButton(text="📅 ۱ ماه", callback_data=f"history_1m_{target_user_id}"),
            InlineKeyboardButton(text="📅 ۳ ماه", callback_data=f"history_3m_{target_user_id}"),
            InlineKeyboardButton(text="📅 ۶ ماه", callback_data=f"history_6m_{target_user_id}"),
        ],
        [InlineKeyboardButton(text="🔙 بازگشت", callback_data=f"back_to_profile_{target_user_id}")]
    ])


async def get_trade_history(current_user_id: int, target_user_id: int, months: int = 3):
    """دریافت تاریخچه معاملات بین دو کاربر"""
    from_date = datetime.utcnow() - timedelta(days=months * 30)
    
    async with AsyncSessionLocal() as session:
        # دریافت کاربر هدف
        target_stmt = select(User).where(User.id == target_user_id)
        target_user = (await session.execute(target_stmt)).scalar_one_or_none()
        
        if not target_user:
            return None, []
        
        # دریافت معاملات بین دو کاربر (یکی لفظ‌دهنده، دیگری پاسخ‌دهنده)
        # فقط با user_id جستجو می‌شود - کاربر جدید به معاملات قبلی دسترسی ندارد
        stmt = (
            select(Trade)
            .options(
                joinedload(Trade.commodity), 
                joinedload(Trade.offer_user),
                joinedload(Trade.responder_user)
            )
            .where(
                and_(
                    Trade.created_at >= from_date,
                    or_(
                        and_(Trade.offer_user_id == current_user_id, Trade.responder_user_id == target_user_id),
                        and_(Trade.offer_user_id == target_user_id, Trade.responder_user_id == current_user_id)
                    )
                )
            )
            .order_by(Trade.created_at.asc())
        )
        result = await session.execute(stmt)
        trades = result.scalars().all()
        
        return target_user, trades


def format_trade_history(trades, target_user, current_user_id: int) -> str:
    """فرمت‌بندی تاریخچه معاملات"""
    if not trades:
        return f"📊 تاریخچه معاملات با {target_user.account_name}\n\n⚠️ معامله‌ای یافت نشد."
    
    text = f"📊 تاریخچه معاملات با {target_user.account_name}\n\n"
    
    for trade in trades[:20]:  # حداکثر 20 معامله
        # تشخیص نوع معامله از دید کاربر فعلی
        if trade.responder_user_id == current_user_id:
            # کاربر فعلی پاسخ‌دهنده بود - trade_type همان نوع عمل اوست
            is_buy = trade.trade_type == TradeType.BUY
        else:
            # کاربر فعلی لفظ‌دهنده بود - عکس trade_type
            is_buy = trade.trade_type != TradeType.BUY
        
        trade_emoji = "🟢" if is_buy else "🔴"
        trade_label = "خرید" if is_buy else "فروش"
        
        # تبدیل به تاریخ شمسی با تایم‌زون ایران
        created_at_iran = trade.created_at.astimezone(IRAN_TZ) if trade.created_at.tzinfo else trade.created_at
        jalali_date = jdatetime.datetime.fromgregorian(datetime=created_at_iran)
        date_str = jalali_date.strftime("%Y/%m/%d")
        
        text += (
            f"{trade_emoji} {trade_label} {trade.commodity.name} "
            f"{trade.quantity} عدد {trade.price:,}\n"
            f"   {date_str}\n\n"
        )
    
    if len(trades) > 20:
        text += f"... و {len(trades) - 20} معامله دیگر"
    
    return text


async def generate_excel(trades, target_user, current_user) -> str:
    """ایجاد فایل Excel با پشتیبانی RTL"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade History"
    ws.sheet_view.rightToLeft = True  # راست به چپ
    
    # هدر - ترتیب RTL (از راست به چپ)
    headers = ["قیمت", "تعداد", "کالا", "نوع", "ساعت", "تاریخ"]
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # داده‌ها - ترتیب RTL
    for row_num, trade in enumerate(trades, 2):
        created_at_iran = trade.created_at.astimezone(IRAN_TZ) if trade.created_at.tzinfo else trade.created_at
        jalali_date = jdatetime.datetime.fromgregorian(datetime=created_at_iran)
        
        # تشخیص نوع معامله از دید کاربر فعلی
        if trade.responder_user_id == current_user.id:
            is_buy = trade.trade_type == TradeType.BUY
        else:
            is_buy = trade.trade_type != TradeType.BUY
        trade_label = "خرید" if is_buy else "فروش"
        
        ws.cell(row=row_num, column=1, value=trade.price)
        ws.cell(row=row_num, column=2, value=trade.quantity)
        ws.cell(row=row_num, column=3, value=trade.commodity.name)
        ws.cell(row=row_num, column=4, value=trade_label)
        ws.cell(row=row_num, column=5, value=jalali_date.strftime("%H:%M"))
        ws.cell(row=row_num, column=6, value=jalali_date.strftime("%Y/%m/%d"))
        
        # سطرهای یکی در میان
        if row_num % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        
        # تراز وسط
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
    
    # عرض ستون‌ها - RTL
    ws.column_dimensions['A'].width = 15  # قیمت
    ws.column_dimensions['B'].width = 10  # تعداد
    ws.column_dimensions['C'].width = 15  # کالا
    ws.column_dimensions['D'].width = 10  # نوع
    ws.column_dimensions['E'].width = 8   # ساعت
    ws.column_dimensions['F'].width = 12  # تاریخ
    
    # ذخیره
    filename = tempfile.mktemp(suffix=".xlsx")
    wb.save(filename)
    
    return filename


async def generate_pdf(trades, target_user, current_user) -> str:
    """ایجاد فایل PDF با فونت فارسی و پشتیبانی RTL"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display
    import os
    
    def reshape_persian(text):
        """تبدیل متن فارسی برای نمایش صحیح RTL"""
        if not text:
            return text
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    
    # ثبت فونت فارسی
    font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'fonts', 'Vazir.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('Vazir', font_path))
        persian_font = 'Vazir'
    else:
        persian_font = 'Helvetica'
    
    filename = tempfile.mktemp(suffix=".pdf")
    doc = SimpleDocTemplate(filename, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    
    # عنوان
    title_style = ParagraphStyle(
        'Title',
        fontName=persian_font,
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    title_text = reshape_persian(f"تاریخچه معاملات با {target_user.account_name}")
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # داده‌های جدول - ترتیب RTL (از راست به چپ)
    headers = [
        reshape_persian("قیمت"),
        reshape_persian("تعداد"),
        reshape_persian("کالا"),
        reshape_persian("نوع"),
        reshape_persian("ساعت"),
        reshape_persian("تاریخ")
    ]
    data = [headers]
    
    for trade in trades:
        created_at_iran = trade.created_at.astimezone(IRAN_TZ) if trade.created_at.tzinfo else trade.created_at
        jalali_date = jdatetime.datetime.fromgregorian(datetime=created_at_iran)
        
        # تشخیص نوع معامله از دید کاربر فعلی
        if trade.responder_user_id == current_user.id:
            is_buy = trade.trade_type == TradeType.BUY
        else:
            is_buy = trade.trade_type != TradeType.BUY
        trade_label = reshape_persian("خرید") if is_buy else reshape_persian("فروش")
        
        data.append([
            f"{trade.price:,}",
            str(trade.quantity),
            reshape_persian(trade.commodity.name),
            trade_label,
            jalali_date.strftime("%H:%M"),
            jalali_date.strftime("%Y/%m/%d")
        ])
    
    # ایجاد جدول - RTL
    col_widths = [80, 50, 100, 50, 50, 80]
    table = Table(data, colWidths=col_widths)
    
    # استایل جدول
    style_commands = [
        # هدر
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5282')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), persian_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
    ]
    
    # سطرهای یکی در میان
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#EDF2F7')))
        else:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.white))
    
    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    
    doc.build(elements)
    
    return filename


# --- دکمه تاریخچه در پروفایل ---
@router.callback_query(F.data.startswith("trade_history_"))
async def show_trade_history(callback: types.CallbackQuery, state: FSMContext, user: Optional[User]):
    if not user:
        await callback.answer("لطفاً ابتدا ثبت‌نام کنید.", show_alert=True)
        return
    
    target_user_id = int(callback.data.split("_")[-1])
    
    target_user, trades = await get_trade_history(user.id, target_user_id, months=3)
    
    if not target_user:
        await callback.answer("کاربر یافت نشد!", show_alert=True)
        return
    
    await state.update_data(history_months=3, history_target_id=target_user_id)
    
    text = format_trade_history(trades, target_user, user.id)
    
    await callback.message.edit_text(
        text,
        reply_markup=get_trade_history_keyboard(target_user_id)
    )
    await callback.answer()


# --- فیلتر تاریخ ---
@router.callback_query(F.data.regexp(r"history_\d+m_\d+"))
async def filter_trade_history(callback: types.CallbackQuery, state: FSMContext, user: Optional[User]):
    if not user:
        return
    
    parts = callback.data.split("_")
    months = int(parts[1].replace("m", ""))
    target_user_id = int(parts[2])
    
    target_user, trades = await get_trade_history(user.id, target_user_id, months=months)
    
    if not target_user:
        await callback.answer("کاربر یافت نشد!", show_alert=True)
        return
    
    await state.update_data(history_months=months, history_target_id=target_user_id)
    
    text = format_trade_history(trades, target_user, user.id)
    
    try:
        await callback.message.edit_text(
            text,
            reply_markup=get_trade_history_keyboard(target_user_id)
        )
    except TelegramBadRequest:
        pass  # پیام تغییر نکرده
    await callback.answer()


# --- دانلود Excel ---
@router.callback_query(F.data.startswith("export_excel_"))
async def export_excel(callback: types.CallbackQuery, state: FSMContext, user: Optional[User], bot: Bot):
    if not user:
        return
    
    await callback.answer("⏳ در حال ایجاد فایل Excel...")
    
    data = await state.get_data()
    months = data.get("history_months", 3)
    target_user_id = int(callback.data.split("_")[-1])
    
    target_user, trades = await get_trade_history(user.id, target_user_id, months=months)
    
    if not trades:
        msg = await callback.message.answer("⚠️ معامله‌ای برای دانلود وجود ندارد.")
        return
    
    try:
        filename = await generate_excel(trades, target_user, user)
        
        # ارسال فایل
        doc_msg = await bot.send_document(
            chat_id=callback.message.chat.id,
            document=FSInputFile(filename, filename=f"trade_history_{target_user.account_name}.xlsx"),
            caption=f"📊 تاریخچه معاملات با {target_user.account_name}\n📅 {months} ماه اخیر"
        )
        
        # حذف فایل موقت
        os.remove(filename)
        
    except Exception as e:
        msg = await callback.message.answer(f"❌ خطا در ایجاد فایل: {str(e)}")


# --- دانلود PDF ---
@router.callback_query(F.data.startswith("export_pdf_"))
async def export_pdf(callback: types.CallbackQuery, state: FSMContext, user: Optional[User], bot: Bot):
    if not user:
        return
    
    await callback.answer("⏳ در حال ایجاد فایل PDF...")
    
    data = await state.get_data()
    months = data.get("history_months", 3)
    target_user_id = int(callback.data.split("_")[-1])
    
    target_user, trades = await get_trade_history(user.id, target_user_id, months=months)
    
    if not trades:
        msg = await callback.message.answer("⚠️ معامله‌ای برای دانلود وجود ندارد.")
        return
    
    try:
        filename = await generate_pdf(trades, target_user, user)
        
        # ارسال فایل
        doc_msg = await bot.send_document(
            chat_id=callback.message.chat.id,
            document=FSInputFile(filename, filename=f"trade_history_{target_user.account_name}.pdf"),
            caption=f"📊 تاریخچه معاملات با {target_user.account_name}\n📅 {months} ماه اخیر"
        )
        
        # حذف فایل موقت
        os.remove(filename)
        
    except Exception as e:
        msg = await callback.message.answer(f"❌ خطا در ایجاد فایل: {str(e)}")


# --- بازگشت به پروفایل ---
@router.callback_query(F.data.startswith("back_to_profile_"))
async def back_to_profile(callback: types.CallbackQuery, state: FSMContext, user: Optional[User]):
    if not user:
        return
    
    target_user_id = int(callback.data.split("_")[-1])
    
    async with AsyncSessionLocal() as session:
        stmt = select(User).where(User.id == target_user_id)
        target_user = (await session.execute(stmt)).scalar_one_or_none()
    
    if target_user:
        profile_text = (
            f"👤 پروفایل عمومی\n\n"
            f"🔸 نام کاربری: {target_user.account_name}\n"
            f"📞 شماره تماس: {target_user.mobile_number}\n"
            f"📍 آدرس: {target_user.address or 'ثبت نشده'}"
        )
        
        await callback.message.edit_text(
            profile_text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📊 تاریخچه معاملات", callback_data=f"trade_history_{target_user_id}")]
            ])
        )
    
    await callback.answer()
